from openpyxl import *
from tkinter import *

wb = load_workbook('F:\\Codes\\Bill Generator\\excel.xlsx') 

sheet = wb.active 

def excel():
	sheet.cell(row=1, column=1).value = "Mob: 9887874343"

	
def focus1(event): 
	date.focus_set() 

def focus2(event): 
	serial_no.focus_set() 

def focus3(event): 
	customer_name.focus_set() 

def focus4(event): 
	designation.focus_set() 

def focus5(event): 
	particular1_name.focus_set() 

def focus6(event):
	particular1_qty.focus_set()

def focus7(event):
	particular1_rate.focus_set()

def focus8(event): 
	particular2_name.focus_set() 

def focus9(event):
	particular2_qty.focus_set()

def focus10(event):
	particular2_rate.focus_set()

def focus11(event): 
	particular3_name.focus_set() 

def focus12(event):
	particular3_qty.focus_set()

def focus13(event):
	particular3_rate.focus_set()

def focus14(event): 
	particular4_name.focus_set() 

def focus15(event):
	particular4_qty.focus_set()

def focus16(event):
	particular4_rate.focus_set()

def focus17(event): 
	particular5_name.focus_set() 

def focus18(event):
	particular5_qty.focus_set()

def focus19(event):
	particular5_rate.focus_set()

def clear(): 
	
	date.delete(0, END) 
	serial_no.delete(0, END) 
	customer_name.delete(0, END) 
	designation.delete(0, END) 
	particular1_name.delete(0, END) 
	particular1_qty.delete(0, END) 
	particular1_rate.delete(0, END) 
	particular2_name.delete(0, END) 
	particular2_qty.delete(0, END) 
	particular2_rate.delete(0, END)  
	particular3_name.delete(0, END) 
	particular3_qty.delete(0, END) 
	particular3_rate.delete(0, END)  
	particular4_name.delete(0, END) 
	particular4_qty.delete(0, END) 
	particular4_rate.delete(0, END)  
	particular5_name.delete(0, END) 
	particular5_qty.delete(0, END) 
	particular5_rate.delete(0, END)  

def insert(): 
	if (date.get() == "" and
		serial_no.get() == "" and
		customer_name.get() == "" and
		designation.get() == "" and
		particular1_name.get() == "" and
		particular1_qty.get() == "" and
		particular1_rate.get() == "" and
		particular2_name.get() == "" and
		particular2_qty.get() == "" and
		particular2_rate.get() == "" and
		particular3_name.get() == "" and
		particular3_qty.get() == "" and
		particular3_rate.get() == "" and
		particular4_name.get() == "" and
		particular4_qty.get() == "" and
		particular4_rate.get() == "" and
		particular5_name.get() == "" and
		particular5_qty.get() == "" and
		particular5_rate.get() == ""): 
			
		print("empty input") 

	else: 
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		sheet.cell(row=5, column=6).value = date.get() 
		sheet.cell(row=5, column=2).value = serial_no.get() 
		sheet.cell(row=6, column=2).value = customer_name.get() 
		sheet.cell(row=7, column=2).value = designation.get() 
		sheet.cell(row=10, column=2).value = particular1_name.get() 
		sheet.cell(row=11, column=4).value = particular1_qty.get() 
		sheet.cell(row=11, column=5).value = particular1_rate.get()
		sheet.cell(row=12, column=2).value = particular2_name.get() 
		sheet.cell(row=13, column=4).value = particular2_qty.get() 
		sheet.cell(row=13, column=5).value = particular2_rate.get() 
		sheet.cell(row=14, column=2).value = particular3_name.get() 
		sheet.cell(row=15, column=4).value = particular3_qty.get() 
		sheet.cell(row=15, column=5).value = particular3_rate.get() 
		sheet.cell(row=16, column=2).value = particular4_name.get() 
		sheet.cell(row=17, column=4).value = particular4_qty.get() 
		sheet.cell(row=17, column=5).value = particular4_rate.get()
		sheet.cell(row=18, column=2).value = particular5_name.get() 
		sheet.cell(row=19, column=4).value = particular5_qty.get() 
		sheet.cell(row=19, column=5).value = particular5_rate.get()   

		# save the file 
		wb.save('F:\\Codes\\Bill Generator\\excel.xlsx') 

		date.focus_set() 

		clear() 


# Driver code 
if __name__ == "__main__": 
	
	root = Tk() 

	root.configure(background='light green') 

	root.title("Bill Generator") 

	root.geometry("1000x400") 

	excel()

	heading = Label(root, text="Harshit Computer's", bg="light green")

	heading2= Label(root, text="Bill Generator", bg="light green")

	date = Label(root, text="Date", bg="light green") 

	serial_no = Label(root, text="Serial No", bg="light green") 

	customer_name = Label(root, text="Customer Name", bg="light green") 

	designation = Label(root, text="Designation", bg="light green") 

	particular1_name = Label(root, text="1. Particular Name ", bg="light green") 

	particular1_qty = Label(root, text="Qty", bg="light green") 

	particular1_rate = Label(root, text="Rate", bg="light green") 

	particular2_name = Label(root, text="2. Particular Name ", bg="light green") 

	particular2_qty = Label(root, text="Qty", bg="light green") 

	particular2_rate = Label(root, text="Rate", bg="light green") 

	particular3_name = Label(root, text="3. Particular Name", bg="light green") 

	particular3_qty = Label(root, text="Qty", bg="light green") 

	particular3_rate = Label(root, text="Rate", bg="light green") 

	particular4_name = Label(root, text="4. Particular Name", bg="light green") 

	particular4_qty = Label(root, text="Qty", bg="light green") 

	particular4_rate = Label(root, text="Rate", bg="light green") 

	particular5_name = Label(root, text="5. Particular Name", bg="light green") 

	particular5_qty = Label(root, text="Qty", bg="light green") 

	particular5_rate = Label(root, text="Rate", bg="light green") 

	heading.grid(row=10,column=19)
	heading2.grid(row=11, column=19)
	date.grid(row=12, column=18) 
	serial_no.grid(row=12, column=21) 
	customer_name.grid(row=13, column=18) 
	designation.grid(row=14, column=18) 
	particular1_name.grid(row=15, column=18) 
	particular1_qty.grid(row=15, column=21) 
	particular1_rate.grid(row=15, column=23)
	particular2_name.grid(row=16, column=18) 
	particular2_qty.grid(row=16, column=21) 
	particular2_rate.grid(row=16, column=23) 
	particular3_name.grid(row=17, column=18) 
	particular3_qty.grid(row=17, column=21) 
	particular3_rate.grid(row=17, column=23) 
	particular4_name.grid(row=18, column=18) 
	particular4_qty.grid(row=18, column=21) 
	particular4_rate.grid(row=18, column=23) 
	particular5_name.grid(row=19, column=18) 
	particular5_qty.grid(row=19, column=21) 
	particular5_rate.grid(row=19, column=23)  

	date = Entry(root) 
	serial_no = Entry(root) 
	customer_name = Entry(root) 
	designation = Entry(root) 
	particular1_name = Entry(root) 
	particular1_qty = Entry(root) 
	particular1_rate = Entry(root) 
	particular2_name = Entry(root) 
	particular2_qty = Entry(root) 
	particular2_rate = Entry(root) 
	particular3_name = Entry(root) 
	particular3_qty = Entry(root) 
	particular3_rate = Entry(root) 
	particular4_name = Entry(root) 
	particular4_qty = Entry(root) 
	particular4_rate = Entry(root) 
	particular5_name = Entry(root) 
	particular5_qty = Entry(root) 
	particular5_rate = Entry(root) 

	date.bind("<Return>", focus1) 

	serial_no.bind("<Return>", focus2) 

	customer_name.bind("<Return>", focus3) 

	designation.bind("<Return>", focus4) 

	particular1_name.bind("<Return>", focus5) 

	particular1_qty.bind("<Return>", focus6) 

	particular1_rate.bind("<Return>", focus7)

	particular2_name.bind("<Return>", focus8) 

	particular2_qty.bind("<Return>", focus9) 

	particular2_rate.bind("<Return>", focus10)

	particular3_name.bind("<Return>", focus11) 

	particular3_qty.bind("<Return>", focus12) 

	particular3_rate.bind("<Return>", focus13)

	particular4_name.bind("<Return>", focus14) 

	particular4_qty.bind("<Return>", focus15) 

	particular4_rate.bind("<Return>", focus16)

	particular5_name.bind("<Return>", focus17) 

	particular5_qty.bind("<Return>", focus18) 

	particular5_rate.bind("<Return>", focus19)

	date.grid(row=12, column=19, ipadx="200") 
	serial_no.grid(row=12, column=22, ipadx="1") 
	customer_name.grid(row=13, column=19, ipadx="200") 
	designation.grid(row=14, column=19, ipadx="200") 
	particular1_name.grid(row=15, column=19, ipadx="200") 
	particular1_qty.grid(row=15, column=22, ipadx="1") 
	particular1_rate.grid(row=15, column=24, ipadx="1")
	particular2_name.grid(row=16, column=19, ipadx="200") 
	particular2_qty.grid(row=16, column=22, ipadx="1") 
	particular2_rate.grid(row=16, column=24, ipadx="1")
	particular3_name.grid(row=17, column=19, ipadx="200") 
	particular3_qty.grid(row=17, column=22, ipadx="1") 
	particular3_rate.grid(row=17, column=24, ipadx="1")
	particular4_name.grid(row=18, column=19, ipadx="200") 
	particular4_qty.grid(row=18, column=22, ipadx="1") 
	particular4_rate.grid(row=18, column=24, ipadx="1")
	particular5_name.grid(row=19, column=19, ipadx="200") 
	particular5_qty.grid(row=19, column=22, ipadx="1") 
	particular5_rate.grid(row=19, column=24, ipadx="1") 
 
	excel()

	submit = Button(root, text="Submit", fg="Black", 
							bg="Red", command=insert) 
	submit.grid(row=30, column=19) 

	# start the GUI 
	root.mainloop() 
